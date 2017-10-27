import os
from copy import copy
from openpyxl import Workbook as wb
from openpyxl import load_workbook as load_wb
from openpyxl.styles import colors, Border, Side, Alignment, Protection, Font, Color

#DATOS DE LOS MODELOS UNAVCO EN EXCEL
##LISTA           Evel Nvel   Modelo de movimiento 
ITRF2008       = ['D11','E11',7,       "ITRF2008"] #Modelo Geodesico +      
ITRF2000AS     = ['D23','E23',8,     "ITRF2000AS"] #Modelo Geodesico + 
ITRF2000DA     = ['D29','E29',9,     "ITRF2000DA"] #Modelo Geodesico +
APKIM2005_DGFI = ['D13','E13',10,"APKIM2005_DGFI"] #Modelo Geodesico +
APKIM2005_IGN  = ['D15','E15',11, "APKIM2005_IGN"] #Modelo Geodesico +
APKIM2000      = ['D27','E27',12,     "APKIM2000"] #Modelo Geodesico
CGPS2004       = ['D19','E19',13,      "CGPS2004"] #Modelo Geodesico
REVEL2000      = ['D21','E21',14,     "REVEL2000"] #Modelo Geodesico
GEODVEL2010    = ['D7','E7',15,     "GEODVEL2010"] #Modelo Geodesico
NNR_MORVEL     = ['D5','E5',      "NNR_MORVEL"] #Modelo Geofisico
HS3_NUVEL1A    = ['D25','E25',   "HS3_NUVEL1A"] #Modelo Geofisico
HS2_NUVEL1A    = ['D31','E31',   "HS2_NUVEL1A"] #Modelo Geofisico
NUVEL1A        = ['D33','E33',       "NUVEL1A"] #Modelo Geofisico
NUVEL1         = ['D35','E35',        "NUVEL1"] #Modelo Geofisico
GSMR2_1        = ['D3','E3',         "GSMR2_1"] #Modelo Combinado
GSMR1_2        = ['D17','E17',       "GSMR1_2"] #Modelo Combinado
MORVEL2010     = ['D9','E9',      "MORVEL2010"] #Modelo Combinado

#Modelo Matricial para seleccion de modelo de movmiento de placa  y seleccion de sus celdas correspodientes en excel

path_vector = 'F:/Archivos y datos GNSS/VECTORES/EXCEL/SA(NNR)/'
path_mmodel = 'D:/MODEL MOTION EXCEL/ESTACIONES SA(NNR)/MAGNAECO/'
path_out = 'F:/Archivos y datos GNSS/TABLAS RESIDUALES/SA(NNR)/'

relacion = 'SA(NNR)'
MODEL = CGPS2004

plantilla = load_wb('D:/TABLA VECTORIAL - ITRF2008_SA(NNR).xlsx')                          ##Se carga el archivo excel, el cual es una plantilla, los archivos base ppp y sus coordenadas
sheetname = str(plantilla.get_sheet_names())                                               ##Comando para obtener el nombre de las hojas de calculo del archivo y convertirlo en string
sheetname = sheetname[2:-2]                                                                ##Se ajusta el nombre del sheetname ya que viene con estos caracteres de mas ['sheetname']
print(sheetname)
sheet_ranges = plantilla[sheetname]                                                        ##Se selecciona la hoja de calculo a trabajar

#creacion de libro de trabajo 
vector_book = wb()                                                                         # Crea el objeto libro de la clase workbook de openpyxl
Vector_sheet = vector_book.create_sheet(MODEL[3], 0)                                     # crea el objeto hoja de la clase hoja                                                                                   # Asigna un string al atributo titulo del objeto hoja
##Copiado de datos base de la plantilla 
for x in range(1,53+1):
    for y in range(1,9+1):
        Vector_sheet.cell(row=x, column=y).value = sheet_ranges.cell(row=x, column=y).value
##Grabado de archivos
Vector_sheet['F1'].value = MODEL[3]

counter = 3
#print(motion_model_book.get_sheet_names())
for file in os.listdir(path_mmodel):
    motion_model_book = load_wb(path_mmodel + file)
    sheetname = str(motion_model_book.get_sheet_names())                                                                              ##Comando para obtener el nombre de las hojas de calculo del archivo y convertirlo en string
    sheetname = sheetname[2:-2]      
    sheet_ranges = motion_model_book[sheetname]                                                                                  ##Se ajusta el nombre del sheetname ya que viene con estos caracteres de mas ['sheetname']
    Vector_sheet.cell(row=counter, column=6).value = sheet_ranges[MODEL[0]].value    
    Vector_sheet.cell(row=counter, column=7).value = sheet_ranges[MODEL[1]].value  
    counter = counter + 1
    print(sheetname)


counter = 3
#print(motion_model_book.get_sheet_names())
for file in os.listdir(path_vector):
    vector_model = load_wb(path_vector + file)
    sheetname = str(vector_model.get_sheet_names())                                                                              ##Comando para obtener el nombre de las hojas de calculo del archivo y convertirlo en string
    sheetname = sheetname[2:6]      
    sheet_ranges = vector_model[sheetname]                                                                                  ##Se ajusta el nombre del sheetname ya que viene con estos caracteres de mas ['sheetname']
    Vector_sheet.cell(row=counter, column=8).value = round(sheet_ranges.cell(row=MODEL[2], column=3).value,2)
    Vector_sheet.cell(row=counter, column=9).value = round(sheet_ranges.cell(row=MODEL[2], column=4).value,2)
    counter = counter + 1
    print(file)

##Guardado de la tabla vectorial
vector_book.save(path_out + MODEL[3] + '_' + relacion + '.xlsx')